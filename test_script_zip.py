import zipfile
from path_config import target_archive
from pypdf import PdfReader
import csv
from openpyxl import load_workbook


def test_add_files_to_zip(add_files_to_zip):
    with zipfile.ZipFile(target_archive, mode='a') as zf:
        for file in zf.namelist():
            print(file)


def test_open_pdf(add_files_to_zip):
    with zipfile.ZipFile(target_archive) as zip_file:
        with zip_file.open('test.pdf') as pdf_file:
            reader = PdfReader(pdf_file)
            assert "VP of Product, Uprising Technology" in reader.pages[5].extract_text()
            # print(reader.pages[5].extract_text())


def test_open_csv(add_files_to_zip):
    with zipfile.ZipFile(target_archive) as zip_file:
        with zip_file.open('test.csv') as csv_file:
            content = csv_file.read().decode('utf-8-sig')
            csvreader = list(csv.reader(content.splitlines()))
            assert ['1;2;3'] in csvreader


def test_open_xlsx(add_files_to_zip):
    with zipfile.ZipFile(target_archive) as zip_file:
        with zip_file.open('test.xlsx') as xlsx_file:
            workbook = load_workbook(xlsx_file).active
            sheet = workbook
            assert sheet.cell(row=1, column=1).value == 1
